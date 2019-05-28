#!/usr/bin/env python3

import openpyxl, json, sys, argparse, os
sys.path.insert(0, './Item')

ap = argparse.ArgumentParser()
ap.add_argument("-f", "--func", required = False, help = "Function name")
args = vars(ap.parse_args())

from functools import reduce
from PrismEncodedItem import *
from PositionEncodedItem import *
from OffsetItem import *
from PrismLookupTable import *

class PrismHelper:
	def __init__(self, items = []):
		self.primeArray = PRIME_ARRAY
		self.primeLength = PRIME_LENGTH
		self.dataRootPath = '../../data/lite'
		self.courseGradeMap = 'CourseGradeMap.json'
		self.horizontalResource = 'CourseGradeEncodedHorizontal.json'
		self.horizontalResourceTest = 'CourseGradeEncodedHorizontal2.json'
		self.excelResourceFile = 'CourseGradeEncoded.xlsx'
		self.excelResourceFile2 = 'CourseGradeEncoded2.xlsx'
		self.items = items
		self.debugMode = False
	# --

	def convertHorizontalRecord(self, resourcePath, outputPath, ouputFileName = 'CourseGradeEncodedHorizontal.json', cacheEncodeVal = False):
		if resourcePath == None:
			resourcePath = self.dataRootPath + '/' + self.excelResourceFile
		# -

		wb = openpyxl.load_workbook(resourcePath)
		ws = wb.active
		wsRange = ws['A{}:I{}'.format(ws.min_row, ws.max_row)]

		studentIDIdx = 1
		semesterIdx = 3
		gradeEncodeIDIdx = 8
		nextSeqSyntax = "->"
		nextPosSyntax = "."
		seqList = []
		encodeValArr = []
		_encodeValArrTemp = []

		curStudentID = "" # Append new string to seqList
		curSemester = 0 # Append new nextSeqSyntax
		curString = "" # To be added into seqList's element

		firstRow = True

		for row in wsRange:
			studentID = str(row[studentIDIdx].value)
			semester = row[semesterIdx].value
			gradeEncodeVal = row[gradeEncodeIDIdx].value
			encodeVal = str(gradeEncodeVal)

			_encodeValArrTemp.append(gradeEncodeVal)
			encodeValArr.append(gradeEncodeVal)

			if firstRow:
				firstRow = False
				curStudentID = studentID
				curSemester = semester
			# -

			# New student
			if studentID != curStudentID:
				_encodeValArrTemp.sort()
				stringTemp = str(reduce(lambda ret, x: '{}.'.format(ret) + str(x), _encodeValArrTemp))

				curString += ((nextSeqSyntax if len(curString) > 0 else '') + stringTemp)
				seqList.append(curString)

				curStudentID = studentID
				curSemester = semester
				curString = ""
				_encodeValArrTemp = []

			# New semester
			elif semester != curSemester:
				curSemester = semester

				_encodeValArrTemp.sort()
				stringTemp = str(reduce(lambda ret, x: '{}.'.format(ret) + str(x), _encodeValArrTemp))

				curString += ((nextSeqSyntax if len(curString) > 0 else '') + stringTemp)
				_encodeValArrTemp = []
			# -
		# -

		# Append last student
		_encodeValArrTemp.sort()
		stringTemp = str(reduce(lambda ret, x: '{}.'.format(ret) + str(x), _encodeValArrTemp))

		curString += ((nextSeqSyntax if len(curString) > 0 else '') + stringTemp)
		seqList.append(curString)
		# -

		fileOutputPath = '{}/{}'.format(outputPath, ouputFileName)
		with open(fileOutputPath, 'w') as fp:
			json.dump(seqList, fp, indent=2)
		# -

		_encodeValArrSet = list(set(encodeValArr))
		_encodeValArrStr = list(map(lambda x: str(x), _encodeValArrSet))
		encodedValPath = None
		if cacheEncodeVal:
			encodedValPath = '{}/{}'.format(outputPath, 'CourseGradeEncodedVal.json')
			with open(encodedValPath, 'w') as fp:
				json.dump(_encodeValArrStr, fp)
		# -

		# print('Number of seqs: {}'.format(len(seqList)))
		# print('Number of items: {}'.format(len(_encodeValArrStr)))
		return fileOutputPath, encodedValPath
	# --


	def createFullPrimalEncoded(self, horizontalRecordPath):
		with open(horizontalRecordPath, 'r') as fp:
			seqList = json.load(fp)

		return self.createFullPrimalEncodedFromData(seqList)
	# --

	def createFullPrimalEncodedFromData(self, data):
		ret = map(lambda item: self.createPrimalItem(item, data), self.items)
		return ret
	# --

	def createPrimalItem(self, item, seqList):
		fullSeqPrimalBlocks = self.createSeqPrimalEncoded(item, seqList)
		fullPosPrimalBlocks = list(map(lambda x: self.createPosPrimalEncoded(item, x), seqList))
		offsets, posItems = self.createPosPrimalEncodedInfo(fullPosPrimalBlocks)
		prismItem = PrismEncodedItem(fullSeqPrimalBlocks, offsets, posItems)
		return prismItem
	# --


	def createPosPrimalEncoded(self, item, sequence):
		itemsetList = sequence.split("->")
		itemsetListLength = len(itemsetList)

		ret = []
		counter = 0

		while counter < itemsetListLength:
			primalVal = 1
			primseIdx = 0

			endIdx = min(itemsetListLength - counter, self.primeLength)
			subItemsetList = itemsetList[counter: counter+endIdx]

			for itemset in subItemsetList:
				if self.doesItemInItemset(item, itemset):
					primalVal *= self.primeArray[primseIdx]
				# -
				primseIdx += 1
			# -
			if primalVal > 1:
				ret.append(primalVal)
			primseIdx = 0
			primalVal = 1
			counter += self.primeLength
		# -
		return ret
	# --

	def createPosPrimalEncodedInfo(self, primalBlocksList):
		offsets = [] # For all seq blocks
		offset = [] # For each seq block
		posItems = []
		curOffsetIdx = 0

		primeArrayCounter = 0 # Start index
		primalBlocksListLength = len(primalBlocksList)

		def appendItemAndOffset(_posItems, _offset, primalBlocks, length, primeVal):
			previous = None
			length = len(primalBlocks)

			for idx in range(0, length):
				item = PositionEncodedItem(primalBlocks[idx], idx, None)
				posItems.append(item)
				if previous != None:
					previous.nextPos = item
				# -
				previous = item
			# -
			_offset.append(OffsetItem(curOffsetIdx, length, primeVal))
		# -

		for idx in range(0, primalBlocksListLength):
			primalBlocks = primalBlocksList[idx] # Single sequence
			primeArrayCounter += 1
			length = len(primalBlocks)
			
			if length > 0: # Ignore length is empty
				appendItemAndOffset(posItems, offset, primalBlocks, length, self.primeArray[primeArrayCounter-1])

			if primeArrayCounter == self.primeLength or idx == primalBlocksListLength - 1:
				offsets.append(offset)
				offset = []
				primeArrayCounter = 0
			# -

			if length > 0:
				curOffsetIdx += length
		# -

		return offsets, posItems
	# --


	def createSeqPrimalEncoded(self, item, seqList):
		seqListLength = len(seqList)
		ret = []
		counter = 0

		while counter < seqListLength:
			primalVal = 1
			primeIdx = 0

			endIdx = min(seqListLength - counter, self.primeLength)
			subSeqList = seqList[counter: counter+endIdx]

			for seq in subSeqList:
				if self.doesItemInSequence(item, seq):
					primalVal *= self.primeArray[primeIdx]
				primeIdx += 1
			# -
			ret.append(primalVal)
			counter += self.primeLength
			primalVal = 1
			primeIdx = 0
		# -
		return ret
	# --


	def doesItemInSequence(self, targetItem, sequence):
		itemsetList = sequence.split('->')
		isExist = len([itemset for itemset in itemsetList if self.doesItemInItemset(targetItem, itemset)]) > 0
		return isExist
	# --


	def doesItemInItemset(self, targetItem, itemset):
		itemList = itemset.split(".")
		isExist = len([item for item in itemList if item == targetItem]) > 0
		return isExist
	# --

	def getPosItemsStr(self, posItems):
		info = ''
		if len(posItems) > 0:
			info = reduce(lambda ret, posItemStr: '{}\n'.format(ret) + posItemStr , list(map(lambda x: x.getDescription(), posItems)))
		return info
	# --


	def getOffsetsStr(self, offsets):
		ret = ""
		if len(offsets) == 0:
			ret = "[]"
		else:
			ret = reduce(lambda ret, itemStr: '{}, '.format(ret) + itemStr, list(map(lambda x: x.getDescription(), offsets)))
		# -
		return ret
	# --

	def getOffsetsListStr(self, offsestList):
		offsetStr = '['
		for offsets in offsestList:
			if len(offsets) == 0: 
				offsetStr += '[], '
			else:
				offsetStr += self.getOffsetsStr(offsets)
			#- 
		offsetStr = offsetStr + ']'
		return offsetStr

	def getPositionStringWithPrime(self, value):
		info = map(lambda x: '1' if value % x == 0 else '0', self.primeArray)
		infoStr = ' '.join(info)
		return infoStr
	# --

	def getPositionArrayToLoopWithPrime(self, value):
		info = list(map(lambda x: 1 if value % x == 0 else 0, self.primeArray))
		return info 
	# --

	def getNextItemFromCurrent(self, item):
		index = self.getNextIndexOfItem(item)
		return None if index == None else self.items[index]
	# --

	def getNextIndexOfItem(self, item):
		try:
			_index = self.items.index(item)
			_index += 1
			return _index if _index < len(self.items) else None
		except ValueError:
			print("[FATAL ERROR] Not found item: {}".format(item))
			exit(1)
	# --

	def reCollectItem(self, threashold = 180):
		seqsPath = 'output/CourseGradeEncodedHorizontal.json'
		with open(seqsPath, 'r') as fp:
			seqs = json.load(fp)
		
		items = []
		for seq in seqs:
			itemsetList = seq.split('->')
			for itemset in itemsetList:
				_items = itemset.split('.')
				_itemsTemp = list(map(lambda x: int(x), _items))
				items += list(filter(lambda x: x > threashold, _itemsTemp))
		# -
		_items = list(set(items))
		_items.sort()
		print('total number: {}'.format(len(_items)))
		return _items
	# --

	def load(self, defaultMode=True):
		# Load items
		if self.debugMode:
			self.items = ['a', 'b', 'c']
		elif defaultMode:
			itemsPath = 'output/CourseGradeEncodedVal_origin.json'
			with open(itemsPath, 'r') as fp:
				self.items = json.load(fp)
		else:
			for idx in range(1, 442):
				self.items.append('{}'.format(idx))
		# -

		# Load resource to be mined
		if defaultMode:
			seqsPath = 'output/CourseGradeEncodedHorizontal_origin.json'
			with open(seqsPath, 'r') as fp:
				self.data = json.load(fp)
		else:
			encodedHorizontalPath = self.dataRootPath + '/' + (self.horizontalResourceTest if self.debugMode else self.horizontalResource)
			with open(encodedHorizontalPath, 'r') as fp:
				self.data = json.load(fp)
		# -

		return self.items, self.data
	# --

	def loadCourseGradeMap(self):
		# Should optimize this process
		path = self.dataRootPath + '/'+ self.courseGradeMap
		with open(path, 'r') as fp:
			self.courseGradeDict = json.load(fp)
		# -
	# --

	def getCourseGradeInfoWithEncodedVal(self, val):
		for key in self.courseGradeDict:
			courseGradeInfo = self.courseGradeDict[key]
			for rangeKey in courseGradeInfo["range"]:
				if courseGradeInfo["range"][rangeKey] == val:
					info = '{} ({})'.format(courseGradeInfo["name"], rangeKey)
					# print('Detected {}'.format(info))
					return info
				# -
			# -
		# -
		print('Not found {}'.format(val))
		exit(1)
		return None
	# --

	def getEncodedValFromCourseGradeInfo(self, info): # info = { 'name', 'range' }
		self.loadCourseGradeMap()
		for key in self.courseGradeDict:
			courseGradeInfo = self.courseGradeDict[key]
			if info['name'] == courseGradeInfo['name']:
				encodedVal = str(courseGradeInfo['range'][info['range']])
				return encodedVal
			# -
		# -
		print('Not found {}'.format(info))
		exit(1)
		return None
	# --

	def parseReadableSeqsMined(self, seqs):
		self.loadCourseGradeMap()
		ret = []
		for seq in seqs:
			itemsetList = seq.split('->')
			seqStr = ''
			for itemset in itemsetList:
				items = itemset.split('.')
				_items = list(map(lambda x: int(x), items))
				info = map(lambda x: self.getCourseGradeInfoWithEncodedVal(x), _items)
				infoStr = ", ".join(info)
				seqStr += ('->' if len(seqStr) > 0 else '') + infoStr
			# -
			ret.append(seqStr)
		# -
		return ret
	# --

	def parsePredictRawStringToEncoded(self, predictRawString):
		rawItemsetList = predictRawString.split('->')
		ret = ''
		for rawItemset in rawItemsetList:
			rawItems = rawItemset.split('.')
			itemsetStr = ''
			for rawItem in rawItems:
				components = rawItem.split('(')
				name = components[0].strip()
				courseRange = components[1][:-1]
				itemsetStr += ('.' if len(itemsetStr) > 0 else '') + self.getEncodedValFromCourseGradeInfo({'name': name, 'range': courseRange})
			# -
			ret += ('->' if len(ret) > 0 else '') + itemsetStr
		# -
		return ret
	# --

	def find(self, query, seqs):
		ret = []
		for seq in seqs:
			index = seq.find(query)
			if index != -1:
				ret.append(seq)
		return ret
	# --

# --- PrismHelper

def convertHorizontal():
	helper = PrismHelper()
	fileOuputPath, encodedValPath = helper.convertHorizontalRecord(None, "output", cacheEncodeVal=True)
	
	cmd = 'open {} {}'.format(fileOuputPath, encodedValPath)
	os.system(cmd)
# --


if __name__ == "__main__":
	helper = PrismHelper()

	func = args['func']

	if func == 'convertHorizontal':
		convertHorizontal()
	elif func == 'reCollectItem':
		helper.reCollectItem()
	elif func == 'loadCourseGradeMap':
		helper.loadCourseGradeMap()
		# helper.getCourseGradeInfoWithEncodedVal(441)
	elif func == 'parsePredictRawStringToEncoded':
		ret = helper.parsePredictRawStringToEncoded('Nhập môn Tin học (B+)->Toán A2 (F)')
		print(ret)
	else:
		print('Not found this func: {}'.format(func))
	# -

# ---

















# 